from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import os
import re


# Creamos la aplicación Flask
app = Flask(__name__)

# Clave secreta para proteger las sesiones de usuario
app.secret_key = "hellocontact_secret_2026"

# Credenciales de acceso definidas en el código
USUARIO_VALIDO    = "admin"
CONTRASENA_VALIDA = "1234"

# Ruta donde está guardado el archivo Excel
EXCEL_PATH = os.path.join(os.path.dirname(__file__), "contactos.xlsx")

# Columnas del archivo Excel en el orden exacto
HEADERS    = ["Nombre", "Apellido", "Teléfono", "Correo", "Dirección", "Categoría", "Favorito", "Estado"]

# Categorías disponibles para los contactos
CATEGORIAS = ["Familia", "Trabajo", "Amigos", "Otro"]


# Ajusta el ancho de las columnas según el texto que contienen
def ajustar_columnas(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


# Crea el Excel si no existe, le da formato a los encabezados
# y agrega 5 contactos de prueba
def inicializar_excel():
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contactos"
        ws.append(HEADERS)
        
        # Encabezados con fondo azul y texto blanco en negrita
        azul_oscuro_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        texto_blanco_bold = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        
        for col_num in range(1, len(HEADERS) + 1):
            celda = ws.cell(row=1, column=col_num)
            celda.fill = azul_oscuro_fill
            celda.font = texto_blanco_bold
            
        # 5 contactos de prueba
            
        ajustar_columnas(ws)
        wb.save(EXCEL_PATH)



# Abre el Excel y devuelve todos los contactos como lista
# El número de fila se usa como identificador para editar y eliminar
def leer_contactos():
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    contactos = []
    for i, fila in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if fila and (fila[0] is not None or fila[2] is not None):
            contactos.append({
                "fila":      i,
                "nombre":    str(fila[0]).strip() if fila[0] is not None else "",
                "apellido":  str(fila[1]).strip() if fila[1] is not None else "",
                "telefono":  str(fila[2]).strip() if fila[2] is not None else "",
                "correo":    str(fila[3]).strip() if fila[3] is not None else "",
                "direccion": str(fila[4]).strip() if fila[4] is not None else "",
                "categoria": str(fila[5]).strip() if fila[5] is not None else "Otro",
                "favorito":  str(fila[6]).strip() if fila[6] is not None else "No",
                "estado":    str(fila[7]).strip() if len(fila) > 7 and fila[7] is not None else "Desconectado",
            })
    return contactos



# Agrega un contacto nuevo al final del Excel
def guardar_contacto(datos):
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.append([
        datos["nombre"], datos["apellido"], datos["telefono"],
        datos["correo"], datos["direccion"], datos["categoria"], datos["favorito"],
    ])
    ajustar_columnas(ws)
    wb.save(EXCEL_PATH)



# Modifica los datos de un contacto existente en el Excel
def actualizar_contacto(num_fila, datos):
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.cell(row=num_fila, column=1).value = datos["nombre"]
    ws.cell(row=num_fila, column=2).value = datos["apellido"]
    ws.cell(row=num_fila, column=3).value = datos["telefono"]
    ws.cell(row=num_fila, column=4).value = datos["correo"]
    ws.cell(row=num_fila, column=5).value = datos["direccion"]
    ws.cell(row=num_fila, column=6).value = datos["categoria"]
    ws.cell(row=num_fila, column=7).value = datos["favorito"]
    ws.cell(row=num_fila, column=8).value = datos["estado"]
    ajustar_columnas(ws)
    wb.save(EXCEL_PATH)


# Elimina la fila del contacto en el Excel
def eliminar_contacto(num_fila):
    inicializar_excel()
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.delete_rows(num_fila)
    ajustar_columnas(ws)
    wb.save(EXCEL_PATH)



# Valida que los campos obligatorios no estén vacíos,
# que el teléfono tenga 8 dígitos y que el correo tenga formato válido
# Devuelve una lista de errores (vacía si todo está bien)

def validar_datos(nombre, telefono, correo):
    errores = []
    if not nombre.strip():
        errores.append("El nombre es obligatorio.")
    if not telefono.strip():
        errores.append("El teléfono es obligatorio.")
    elif not re.fullmatch(r"\d{8}", telefono.strip()):
        errores.append("El teléfono debe tener exactamente 8 dígitos.")
    if not correo.strip():
        errores.append("El correo es obligatorio.")
    elif not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", correo.strip()):
        errores.append("El correo electrónico no tiene un formato válido.")
    return errores


# Decorador que protege las rutas
# Si el usuario no ha iniciado sesión, lo manda al login
def login_requerido(f):
    from functools import wraps
    @wraps(f)
    def decorado(*args, **kwargs):
        if "usuario" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorado


# Pantalla de inicio de sesión
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario    = request.form.get("usuario", "")
        contrasena = request.form.get("contrasena", "")
        if usuario == USUARIO_VALIDO and contrasena == CONTRASENA_VALIDA:
            session["usuario"] = usuario
            return redirect(url_for("contactos"))
        else:
            flash("Credenciales incorrectas. Intenta de nuevo.", "error")
    return render_template("login.html")


# Pantalla de registro
@app.route("/registro", methods=["GET", "POST"])
def registro():
    if request.method == "POST":
        flash("Cuenta creada. Usa admin / 1234 para ingresar.", "success")
        return redirect(url_for("login"))
    return render_template("registro.html")


# Cierra la sesión y redirige al login
@app.route("/logout")
def logout():
    session.pop("usuario", None)
    return redirect(url_for("login"))


# Lista todos los contactos, permite ordenar A-Z
@app.route("/contactos")
@login_requerido
def contactos():
    lista = leer_contactos()
    orden = request.args.get("orden", "")
    if orden == "az":
        lista = sorted(lista, key=lambda c: c["nombre"].lower())
    return render_template("contactos.html", contactos=lista, orden=orden)


# Busca contactos por nombre o apellido
@app.route("/buscar")
@login_requerido
def buscar():
    query = request.args.get("q", "").strip().lower()
    resultados = []
    if query:
        resultados = [
            c for c in leer_contactos()
            if query in c["nombre"].lower() or query in c["apellido"].lower()
        ]
    return render_template("buscar.html", resultados=resultados, query=query)


# Formulario para agregar un nuevo contacto
@app.route("/agregar", methods=["GET", "POST"])
@login_requerido
# Si el usuario envió el formulario
def agregar():
    if request.method == "POST":
        # .strip() elimina espacios en blanco al inicio y al final
        nombre    = request.form.get("nombre", "").strip()
        apellido  = request.form.get("apellido", "").strip()
        telefono  = request.form.get("telefono", "").strip()
        correo    = request.form.get("correo", "").strip()
        direccion = request.form.get("direccion", "").strip()
        categoria = request.form.get("categoria", "Otro")
        favorito  = "Sí" if request.form.get("favorito") else "No"
        
        errores  = validar_datos(nombre, telefono, correo)
        # Si hay errores de validación los mostramos y recargamos el formulario
        if errores:
            for e in errores:
                flash(e, "error")
            return render_template("agregar.html", categorias=CATEGORIAS, form=request.form)

        # Si todo está correcto guardamos el contacto en el Excel
        guardar_contacto({
            "nombre": nombre, "apellido": apellido, "telefono": telefono,
            "correo": correo, "direccion": direccion, "categoria": categoria, "favorito": favorito
        })
        # Mostramos mensaje de éxito al usuario
        flash("Contacto agregado exitosamente.", "success")
        return redirect(url_for("contactos"))
    return render_template("agregar.html", categorias=CATEGORIAS, form={})


# Muestra todos los datos de un contacto
@app.route("/detalle/<int:fila>")
@login_requerido
def detalle(fila):
    contactos_lista = leer_contactos()
    contacto = next((c for c in contactos_lista if c["fila"] == fila), None)
    if not contacto:
        flash("Contacto no encontrado.", "error")
        return redirect(url_for("contactos"))
    return render_template("detalle.html", contacto=contacto)


# Formulario para editar un contacto existente
@app.route("/editar/<int:fila>", methods=["GET", "POST"])
@login_requerido
def editar(fila):
    contactos_lista = leer_contactos()
    contacto = next((c for c in contactos_lista if c["fila"] == fila), None)
    # Si no existe el contacto mostramos un error
    if not contacto:
        # Redirigimos a la lista de contactos
        flash("Contacto no encontrado.", "error")
        return redirect(url_for("contactos"))
        
    if request.method == "POST":
        # Leemos los nuevos datos del formulario
        nombre    = request.form.get("nombre", "").strip()
        apellido  = request.form.get("apellido", "").strip()
        telefono  = request.form.get("telefono", "").strip()
        correo    = request.form.get("correo", "").strip()
        direccion = request.form.get("direccion", "").strip()
        categoria = request.form.get("categoria", "Otro")
        favorito  = "Sí" if request.form.get("favorito") else "No"
        
        errores  = validar_datos(nombre, telefono, correo)
        if errores:
            for e in errores:
                flash(e, "error")
            return render_template("editar.html", contacto=contacto, categorias=CATEGORIAS)
            
        actualizar_contacto(fila, {
            "nombre": nombre, "apellido": apellido, "telefono": telefono,
            "correo": correo, "direccion": direccion, "categoria": categoria, "favorito": favorito
        })
        flash("Contacto actualizado correctamente.", "success")
        return redirect(url_for("detalle", fila=fila))
    return render_template("editar.html", contacto=contacto, categorias=CATEGORIAS)


# Elimina un contacto del Excel
@app.route("/eliminar/<int:fila>", methods=["POST"])
@login_requerido
def eliminar(fila):
    eliminar_contacto(fila)
    flash("Contacto eliminado correctamente.", "success")
    return redirect(url_for("contactos"))


# Pantalla de reporte con estadísticas
@app.route("/reporte")
@login_requerido
def reporte():
    lista        = leer_contactos()
    total        = len(lista)
    favoritos    = sum(1 for c in lista if c["favorito"] == "Sí")
    no_favoritos = total - favoritos
    
    por_categoria = {}
    for cat in CATEGORIAS:
        # Contamos cuántos contactos hay en cada categoría
        por_categoria[cat] = sum(1 for c in lista if c["categoria"] == cat)

    # Enviamos los datos al template del reporte
    return render_template("reporte.html",
                           total=total, favoritos=favoritos,
                           no_favoritos=no_favoritos,
                           por_categoria=por_categoria, categorias=CATEGORIAS)


# Descarga el archivo Excel al navegador
@app.route("/exportar")
@login_requerido
def exportar():
    inicializar_excel()
    return send_file(
        EXCEL_PATH,
        as_attachment=True,
        download_name="contactos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    # Creamos el Excel si no existe
    inicializar_excel()
    # Iniciamos el servidor en modo desarrollo
    app.run(debug=True)
